import openpyxl
from openpyxl.styles import Alignment
import os
from pathlib import Path

# Dump locations
dumps = Path(__file__).parent.parent / "dumps"
loc_rightmove = Path(dumps) / "rightmove.xlsx"
loc_openrent = Path(dumps) / "openrent.xlsx"
loc_openmove = Path(dumps) / "rightmove+openrent.xlsx"

# Create dumps folder
if not os.path.isdir(dumps):
    os.makedirs(dumps)

# Choose workbook
if os.path.isfile(loc_rightmove):
    wb = openpyxl.load_workbook(loc_rightmove)
    loc_output = loc_openmove
else:
    wb = openpyxl.Workbook()
    loc_output = loc_openrent

# Choose work sheet
ws = wb.active

# Free variables
COLUMN_START = 2
COLUMN_AMOUNT = 4
COLUMN_POSTCODE = 1

# Insert new columns and shift merged header cells
ws.insert_cols(2, amount=4)
merged_cells = ws.merged_cells.ranges
for cell in merged_cells:
    cell.shift(4, 0)

# Create header
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
header = ws.cell(row=1, column=2)
header.value = "Per Room"
header.alignment = Alignment(horizontal='center')

# Create sub headers
ws.cell(row=2, column=1).value = "Post Code"
ws.cell(row=2, column=2).value = "Total Rent"
ws.cell(row=2, column=3).value = "Total Let"
ws.cell(row=2, column=4).value = "% Let"
ws.cell(row=2, column=5).value = "Average Let Price"


# Insert / Update a row based on a postcode.
def insert(postcode, data):
    found = None
    postcodes = []

    # search for particular postcode
    for i in ws.rows:
        row = i[0].row

        # ignore header and subheader rows
        if row <= 2:
            continue

        # gather existing postcodes
        if i[0].value:
            postcodes.append(i[0].value)

        cell = ws.cell(row=row, column=1)

        # cell found
        if cell.value == postcode:
            found = cell
            break

    # calculate the row position to in
    if found:
        row = found.row
    else:
        postcodes.append(postcode)
        postcodes.sort()
        row = postcodes.index(postcode) + 3

    # create new postcode row and insert empty data on all other columns
    if not found:
        ws.insert_rows(row, amount=1)
        for column in range(1, ws.max_column+1):
            ws.cell(row=row, column=column).value = postcode if column == 1 else 0

    # insert into postcode row with openrent data
    ws.cell(row=row, column=2).value = data["totalRentCount"]
    ws.cell(row=row, column=3).value = data["totalLetCount"]
    ws.cell(row=row, column=4).value = data["totalLetPercentage"]
    ws.cell(row=row, column=5).value = data["totalLetAverage"]


# Save workbook to file
def save():
    # fill up empty columns
    for col in ws.iter_cols(min_col=2, max_col=5):
        for cell in col:
            if cell.value:
                continue
            cell.value = 0

    wb.save(loc_output)
